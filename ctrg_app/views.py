from urllib.parse import urlparse

from django.contrib.auth.hashers import make_password
from django.utils import timezone
from django.db import connection
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect, get_object_or_404
# 1. Add at top with other imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .forms import GrantCycleForm
from .models import SrcChairs, Reviewers, Users, Departments, Grantcycles, Reviewassignments, Stage1Reviews, \
    Stage2Reviews, Proposaldocuments, Schools
from django.shortcuts import render, redirect # <-- Import redirect here
from django.contrib import messages
from .models import Proposals, Departments
from django.db.models import Max
from .models import Proposals, Proposaldocuments, Departments
from .models import Proposaldocuments
from django.http import HttpResponse
from django.core.mail import EmailMessage, get_connection, send_mail
from django.conf import settings
from django.urls import reverse
from .email_delivery import EmailDeliveryError, send_transactional_email
from .cycle_activation import (
    attach_active_cycle_flags,
    get_active_cycle_id_for_school,
    set_active_cycle_for_school,
)
from .stage1_scoring import calculate_stage1_total, ensure_stage1_total
from .p3_services import save_upload_to_media


def build_assignment_email(request, reviewer, proposal, assignment):
    assignment_url = request.build_absolute_uri(
        reverse("respond_to_assignment", args=[assignment.assignment_id])
    )
    subject = "New Proposal Assignment - CTRG System"
    body = (
        f"Dear {reviewer.user.username},\n\n"
        "You have been assigned to review a proposal.\n\n"
        f"Proposal: {proposal.title}\n\n"
        "Please log in and respond to this assignment here:\n"
        f"{assignment_url}\n\n"
        "CTRG System"
    )
    html_body = f"""
        <p>Dear {reviewer.user.username},</p>
        <p>You have been assigned to review a proposal.</p>
        <p><strong>Proposal:</strong> {proposal.title}</p>
        <p>
            Please log in and respond to this assignment here:<br>
            <a href="{assignment_url}">{assignment_url}</a>
        </p>
        <p>CTRG System</p>
    """
    return subject, body, html_body


def send_assignment_email(request, reviewer, proposal, assignment):
    recipient = (getattr(reviewer.user, "email", "") or "").strip()
    if not recipient:
        return False, "Reviewer assignment saved, but the reviewer has no email address."

    subject, body, html_body = build_assignment_email(request, reviewer, proposal, assignment)

    try:
        result = send_transactional_email(
            subject=subject,
            recipient_list=[recipient],
            text_body=body,
            html_body=html_body,
        )
    except EmailDeliveryError as exc:
        return False, f"Reviewer assignment saved, but the email could not be sent: {exc}"

    if result.get("local_only"):
        return (
            False,
            "Reviewer assignment email was generated with a local-only email backend. "
            "Check the runserver terminal or configure a real SMTP/API backend for inbox delivery.",
        )

    if result.get("transport") == "brevo":
        return True, f"Reviewer assignment email sent via Brevo to {recipient}."

    if result.get("transport") == "resend":
        return True, f"Reviewer assignment email sent via Resend to {recipient}."

    return True, f"Reviewer assignment email sent to {recipient}."


def local_now_naive():
    return timezone.localtime(timezone.now()).replace(tzinfo=None)


def parse_local_datetime_input(value):
    """Convert a datetime-local value from the browser into Dhaka-local naive storage."""
    if not value:
        return None

    from django.utils.dateparse import parse_datetime

    if len(value) == 16:
        value = f"{value}:00"

    parsed = parse_datetime(value)
    if parsed is None:
        return None

    if not timezone.is_naive(parsed):
        parsed = timezone.make_naive(parsed, timezone.get_current_timezone())

    return parsed


def format_cycle_datetime_local_input(value):
    """Render stored cycle datetimes back into Asia/Dhaka datetime-local strings."""
    if not value:
        return ""

    if timezone.is_naive(value):
        return value.strftime("%Y-%m-%dT%H:%M")

    return timezone.localtime(value, timezone.get_current_timezone()).strftime("%Y-%m-%dT%H:%M")


def format_cycle_datetime_display(value):
    if not value:
        return "-"

    if timezone.is_naive(value):
        return value.strftime("%b %d, %Y %I:%M %p")

    return timezone.localtime(value, timezone.get_current_timezone()).strftime("%b %d, %Y %I:%M %p")


def attach_cycle_display_fields(cycle):
    cycle.stage1_start_date_local = format_cycle_datetime_local_input(cycle.stage1_start_date)
    cycle.stage1_end_date_local = format_cycle_datetime_local_input(cycle.stage1_end_date)
    cycle.revision_duration_days_local = format_cycle_datetime_local_input(cycle.revision_duration_days)
    cycle.stage2_start_date_local = format_cycle_datetime_local_input(cycle.stage2_start_date)
    cycle.stage2_end_date_local = format_cycle_datetime_local_input(cycle.stage2_end_date)

    cycle.stage1_start_date_display = format_cycle_datetime_display(cycle.stage1_start_date)
    cycle.stage1_end_date_display = format_cycle_datetime_display(cycle.stage1_end_date)
    cycle.revision_duration_days_display = format_cycle_datetime_display(cycle.revision_duration_days)
    cycle.stage2_start_date_display = format_cycle_datetime_display(cycle.stage2_start_date)
    cycle.stage2_end_date_display = format_cycle_datetime_display(cycle.stage2_end_date)
    return cycle




def login_view(request):
    if request.user.is_authenticated:
        return role_redirect(request.user)
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return role_redirect(user)
        else:
            messages.error(request, 'Invalid username or password')
    return render(request, 'login.html')

def signup_view(request):
    departments = Departments.objects.all().order_by('department_name')

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        dept_id = request.POST.get('department') or None
        area_of_expertise = request.POST.get('area_of_expertise') or None
        profile_picture = request.FILES.get('profile_picture')

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect('signup')

        if Users.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('signup')

        try:
            profile_picture_path = None
            if profile_picture:
                profile_picture_path = save_upload_to_media(profile_picture, folder="profiles")

            user = Users.objects.create(
                username=username,
                email=email,
                password=make_password(password),  # important
                department_id=dept_id,
                area_of_expertise=area_of_expertise,
                profile_picture=profile_picture_path,
                is_active=True
            )
            messages.success(request, "Account created successfully! Please login.")
            return redirect('login')
        except Exception as e:
            messages.error(request, f"Error creating user: {e}")
            return redirect('signup')

    return render(request, 'signup.html', {'departments': departments})


def role_redirect(user):
    return redirect('dashboard')

def dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')
    user = request.user
    is_chair = SrcChairs.objects.filter(user=user, is_active=True).exists()
    is_reviewer = Reviewers.objects.filter(user=user, is_active=True).exists()

    context = {
        'is_chair': is_chair,
        'is_reviewer': is_reviewer,
        'is_pi': True,
    }
    return render(request, 'dashboard.html', context)


def logout_view(request):
    logout(request)
    return redirect('login')


# --- SRC CHAIR MANAGEMENT (Person 1 Tasks) ---

def chair_dashboard(request):
    # Security: Ensure only SRC Chair can access
    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    # Context data for the command center
    cycles = list(
        Grantcycles.objects.filter(school=chair.school).select_related('school').order_by('-year', '-cycle_id')
    )
    active_cycle_id = attach_active_cycle_flags(cycles, chair.school_id)

    context = {
        'cycles': cycles,
        'reviewer_count': Reviewers.objects.filter(is_active=True, department__school=chair.school).count(),
        'chair_school': chair.school,
        'active_cycle_id': active_cycle_id,
        # Add proposal stats here once Person 3 creates the models
    }
    return render(request, 'chair/dashboard.html', context)


def create_grant_cycle(request):
    # Security check: Only active SRC Chair can create cycles
    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    if chair.school_id is None:
        messages.error(request, 'Your SRC Chair account is not assigned to a school.')
        return redirect('chair_dashboard')

    if request.method == 'POST':
        form = GrantCycleForm(request.POST)
        if form.is_valid():
            cleaned = form.cleaned_data
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO grantcycles (
                        created_by_src_id,
                        cycle_name,
                        year,
                        stage1_start_date,
                        stage1_end_date,
                        revision_duration_days,
                        stage2_start_date,
                        stage2_end_date,
                        acceptance_threshold,
                        max_reviewers_per_proposal,
                        school_id
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        chair.src_id,
                        cleaned.get('cycle_name'),
                        cleaned.get('year'),
                        cleaned.get('stage1_start_date'),
                        cleaned.get('stage1_end_date'),
                        cleaned.get('revision_duration_days'),
                        cleaned.get('stage2_start_date'),
                        cleaned.get('stage2_end_date'),
                        cleaned.get('acceptance_threshold'),
                        cleaned.get('max_reviewers_per_proposal'),
                        chair.school_id,
                    ],
                )
            if not get_active_cycle_id_for_school(chair.school_id):
                newest_cycle = (
                    Grantcycles.objects
                    .filter(school=chair.school)
                    .order_by('-cycle_id')
                    .first()
                )
                if newest_cycle:
                    set_active_cycle_for_school(chair.school_id, newest_cycle.cycle_id)
            messages.success(request, 'Grant Cycle created successfully!')
            return redirect('chair_dashboard')
        messages.error(request, 'Please correct the grant cycle form and try again.')
    else:
        form = GrantCycleForm()
    return render(request, 'chair/create_cycle.html', {
        'form': form,
        'chair_school': chair.school,
    })


# --- REVIEWER MANAGEMENT ---

def reviewer_list(request):
    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    # List all reviewers for resource management
    reviewers = (
        Reviewers.objects
        .filter(is_active=True, department__school=chair.school)
        .select_related('user', 'department', 'department__school')
    )
    return render(request, 'chair/reviewer_list.html', {
        'reviewers': reviewers,
        'chair_school': chair.school,
    })


def email_reviewers(request):
    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    reviewers = list(
        Reviewers.objects
        .filter(is_active=True, department__school=chair.school)
        .select_related('user', 'department', 'department__school')
        .order_by('user__username')
    )

    recipient_map = {}
    missing_email_count = 0
    for reviewer in reviewers:
        email = (getattr(reviewer.user, 'email', '') or '').strip()
        if not email:
            missing_email_count += 1
            continue
        if email not in recipient_map:
            recipient_map[email] = getattr(reviewer.user, 'username', 'Reviewer')

    subject = "CTRG notice from the SRC Chair"
    message_body = ""

    if request.method == 'POST':
        subject = (request.POST.get('subject') or '').strip()
        message_body = (request.POST.get('message') or '').strip()

        if not subject or not message_body:
            messages.error(request, 'Subject and message are required.')
        elif not recipient_map:
            messages.error(request, 'No active reviewers with email addresses were found for your school.')
        else:
            sent_count = 0
            failed_count = 0
            local_only_count = 0

            for recipient, reviewer_name in recipient_map.items():
                body = (
                    f"Dear {reviewer_name},\n\n"
                    f"{message_body}\n\n"
                    "CTRG System"
                )
                html_body = f"""
                    <p>Dear {reviewer_name},</p>
                    <p>{message_body.replace(chr(10), '<br>')}</p>
                    <p>CTRG System</p>
                """

                try:
                    result = send_transactional_email(
                        subject=subject,
                        text_body=body,
                        recipient_list=[recipient],
                        html_body=html_body,
                    )
                except EmailDeliveryError:
                    failed_count += 1
                    continue

                sent_count += 1
                if result.get("local_only"):
                    local_only_count += 1

            if sent_count:
                messages.success(request, f'Email sent to {sent_count} reviewer(s).')
                subject = "CTRG notice from the SRC Chair"
                message_body = ""
            if missing_email_count:
                messages.warning(request, f'{missing_email_count} reviewer(s) were skipped because no email address is set.')
            if failed_count:
                messages.error(request, f'{failed_count} email(s) could not be sent.')
            if local_only_count:
                messages.warning(
                    request,
                    f'{local_only_count} email(s) used a local-only backend and were not delivered to inboxes.',
                )

    return render(request, 'chair/email_reviewers.html', {
        'chair_school': chair.school,
        'reviewer_count': len(reviewers),
        'recipient_count': len(recipient_map),
        'missing_email_count': missing_email_count,
        'subject': subject,
        'message_body': message_body,
    })


def reviewer_detail(request, reviewer_id):
    if not SrcChairs.objects.filter(user=request.user, is_active=True).exists():
        return redirect('login')

    reviewer = get_object_or_404(
        Reviewers.objects.select_related('user', 'department', 'department__school'),
        reviewer_id=reviewer_id
    )
    profile_picture = (getattr(reviewer.user, 'profile_picture', '') or '').strip()
    profile_picture_url = ''
    if profile_picture:
        parsed_profile = urlparse(profile_picture)
        if parsed_profile.scheme and parsed_profile.netloc:
            profile_picture_url = profile_picture
        elif profile_picture.startswith(settings.MEDIA_URL):
            profile_picture_url = profile_picture
        else:
            profile_picture_url = f"{settings.MEDIA_URL}{profile_picture.lstrip('/')}"

    return render(request, 'chair/reviewer_detail.html', {
        'reviewer': reviewer,
        'profile_picture_url': profile_picture_url,
    })


def create_reviewer(request):
    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    # Exclude only users who already have an active reviewer assignment.
    existing_reviewer_user_ids = Reviewers.objects.filter(is_active=True).values_list('user_id', flat=True)
    available_users = (
        Users.objects
        .filter(department__school=chair.school)
        .exclude(user_id__in=existing_reviewer_user_ids)
        .select_related('department', 'department__school')
        .order_by('username')
    )

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        max_load = request.POST.get('max_load')

        try:
            selected_user = Users.objects.select_related('department', 'department__school').get(
                user_id=user_id,
                department__school=chair.school
            )
            if selected_user.department_id is None:
                messages.error(request, 'The selected user does not have a department assigned.')
                return redirect('create_reviewer')

            existing_reviewer = Reviewers.objects.filter(user=selected_user).order_by('-reviewer_id').first()
            if existing_reviewer:
                existing_reviewer.department = selected_user.department
                existing_reviewer.max_review_load = max_load
                existing_reviewer.start_date = timezone.now().date()
                existing_reviewer.end_date = None
                existing_reviewer.is_active = True
                existing_reviewer.save()
                messages.success(request, 'Reviewer reactivated successfully!')
            else:
                Reviewers.objects.create(
                    user=selected_user,
                    department=selected_user.department,
                    max_review_load=max_load,
                    start_date=timezone.now().date(),
                    is_active=True
                )
                messages.success(request, 'Reviewer assigned successfully!')
            return redirect('reviewer_list')
        except Users.DoesNotExist:
            messages.error(request, 'Selected user is not valid for your school.')
            return redirect('create_reviewer')
        except Exception as e:
            messages.error(request, f"Error: {e}")

    return render(request, 'chair/create_reviewer.html', {
        'available_users': available_users,
        'chair_school': chair.school,
    })

# --- DASHBOARD PLACEHOLDERS ---

def reviewer_dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')

    reviewer = Reviewers.objects.filter(user=request.user, is_active=True).first()
    if not reviewer:
        return redirect('login')

    assignments = (
        Reviewassignments.objects
        .filter(reviewer=reviewer, is_active=True)
        .select_related("proposal", "proposal__department")
    )

    assignment_ids = [assignment.assignment_id for assignment in assignments]
    stage1_assignment_ids = set(
        Stage1Reviews.objects.filter(assignment_id__in=assignment_ids).values_list("assignment_id", flat=True)
    )
    stage2_assignment_ids = set(
        Stage2Reviews.objects.filter(assignment_id__in=assignment_ids).values_list("assignment_id", flat=True)
    )

    # Use naive project-local time so comparisons match the configured Django timezone.
    now = local_now_naive()

    def _naive(d):
        """Strip timezone from a stored datetime so it compares with naive local time."""
        if d is None:
            return None
        return d.replace(tzinfo=None) if getattr(d, 'tzinfo', None) else d

    for assignment in assignments:
        assignment.has_stage1_review = assignment.assignment_id in stage1_assignment_ids
        assignment.has_stage2_review = assignment.assignment_id in stage2_assignment_ids
        cycle = assignment.proposal.cycle
        # Stage 1 window
        s1_start = _naive(cycle.stage1_start_date)
        s1_end   = _naive(cycle.stage1_end_date)
        assignment.stage1_window_open = (
            (s1_start is None or now >= s1_start) and
            (s1_end   is None or now <= s1_end)
        )
        # Stage 2 window
        s2_start = _naive(cycle.stage2_start_date)
        s2_end   = _naive(cycle.stage2_end_date)
        assignment.stage2_window_open = (
            (s2_start is None or now >= s2_start) and
            (s2_end   is None or now <= s2_end)
        )
        # Human-readable labels
        assignment.stage1_window_label = (
            f"Opens {s1_start.strftime('%b %d %H:%M:%S')}" if s1_start and now < s1_start
            else f"Closed {s1_end.strftime('%b %d %H:%M:%S')}" if s1_end and now > s1_end
            else None
        )
        assignment.stage2_window_label = (
            f"Opens {s2_start.strftime('%b %d %H:%M:%S')}" if s2_start and now < s2_start
            else f"Closed {s2_end.strftime('%b %d %H:%M:%S')}" if s2_end and now > s2_end
            else None
        )

    context = {"assignments": assignments, "now": now}
    return render(request, "reviews/reviewer_dashboard.html", context)


def respond_to_assignment(request, assignment_id):
    if not request.user.is_authenticated:
        return redirect('login')

    reviewer = Reviewers.objects.filter(user=request.user, is_active=True).first()
    if not reviewer:
        return redirect('login')

    assignment = Reviewassignments.objects.filter(
        assignment_id=assignment_id,
        reviewer=reviewer
    ).first()

    if not assignment:
        messages.error(request, "Assignment not found.")
        return redirect('reviewer_dashboard')

    if assignment.acceptance_status != 'PENDING':
        messages.warning(request, "You have already responded to this assignment.")
        return redirect('reviewer_dashboard')

    if request.method == "POST":
        response = request.POST.get("response")
        if response == "ACCEPTED":
            assignment.acceptance_status = "ACCEPTED"
            assignment.save()
            messages.success(request, "You have accepted the assignment.")
        elif response == "REJECTED":
            assignment.acceptance_status = "REJECTED"
            assignment.save()
            messages.success(request, "You have rejected the assignment.")
        else:
            messages.error(request, "Invalid response.")

    return redirect('reviewer_dashboard')

def pi_dashboard(request):
    return render(request, 'pi/dashboard.html')

def cycle_list(request):
    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')
    cycles = list(
        Grantcycles.objects.select_related('school').filter(school=chair.school).order_by('-year', '-cycle_id')
    )
    active_cycle_id = attach_active_cycle_flags(cycles, chair.school_id)
    # Add proposal count per cycle
    for cycle in cycles:
        attach_cycle_display_fields(cycle)
        cycle.proposal_count = Proposals.objects.filter(cycle=cycle).count()
    return render(request, 'chair/cycle_list.html', {
        'cycles': cycles,
        'chair_school': chair.school,
        'active_cycle_id': active_cycle_id,
    })

def proposals_by_cycle(request, cycle_id):
    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')
    cycle = get_object_or_404(Grantcycles.objects.select_related('school'), cycle_id=cycle_id, school=chair.school)
    attach_cycle_display_fields(cycle)
    proposals = Proposals.objects.filter(cycle=cycle).select_related('pi_user', 'department__school').order_by('-proposal_id')
    return render(request, 'proposals/proposals_by_cycle.html', {
        'cycle': cycle,
        'proposals': proposals
    })

def edit_reviewer(request, reviewer_id):
    if not SrcChairs.objects.filter(user=request.user, is_active=True).exists():
        return redirect('login')

    try:
        reviewer = Reviewers.objects.get(reviewer_id=reviewer_id)
    except Reviewers.DoesNotExist:
        messages.error(request, 'Reviewer does not exist!')
        return redirect('reviewer_list')
    departments = Departments.objects.all()
    if request.method == 'POST':
        reviewer.department_id = request.POST.get('department')
        reviewer.max_review_load = request.POST.get('max_load')
        reviewer.is_active = request.POST.get('is_active', '1') in {'1', 'true', 'True', 'on'}
        reviewer.save()
        messages.success(request, 'Reviewer assigned successfully!')
        return redirect('reviewer_list')
    context = { "reviewer": reviewer , "departments": departments}
    return render(request, 'chair/edit_reviewer.html', context)

def deactivate_reviewer(request, reviewer_id):
    if not SrcChairs.objects.filter(user=request.user, is_active=True).exists():
        return redirect('login')

    if request.method == 'POST':
        try:
            reviewer = Reviewers.objects.get(reviewer_id=reviewer_id)
            print(f"Found reviewer: {reviewer.user.username}")
            print(f"Current is_active: {reviewer.is_active}")
            reviewer.is_active = False
            reviewer.end_date = timezone.now().date()
            reviewer.save()
            print(f"After save is_active: {reviewer.is_active}")
            messages.success(request, f'Reviewer {reviewer.user.username} has been deactivated.')
        except Reviewers.DoesNotExist:
            messages.error(request, 'Reviewer not found.')
        except Exception as e:
            print(f"ERROR: {e}")
            messages.error(request, f'Error: {e}')

    return redirect('reviewer_list')

def edit_grant_cycle(request, cycle_id):
    try:
        chair = SrcChairs.objects.get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    try:
        cycle = Grantcycles.objects.get(cycle_id=cycle_id)
    except Grantcycles.DoesNotExist:
        messages.error(request, 'Cycle not found.')
        return redirect('cycle_list')

    schools = Schools.objects.all().order_by('school_name')

    if request.method == 'POST':
        form = GrantCycleForm(request.POST, instance=cycle)
        if form.is_valid():
            cleaned = form.cleaned_data
            school = cleaned.get('school')
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE grantcycles
                    SET cycle_name = %s,
                        year = %s,
                        school_id = %s,
                        stage1_start_date = %s,
                        stage1_end_date = %s,
                        revision_duration_days = %s,
                        stage2_start_date = %s,
                        stage2_end_date = %s,
                        acceptance_threshold = %s,
                        max_reviewers_per_proposal = %s
                    WHERE cycle_id = %s
                    """,
                    [
                        cleaned.get('cycle_name'),
                        cleaned.get('year'),
                        school.school_id if school else None,
                        cleaned.get('stage1_start_date'),
                        cleaned.get('stage1_end_date'),
                        cleaned.get('revision_duration_days'),
                        cleaned.get('stage2_start_date'),
                        cleaned.get('stage2_end_date'),
                        cleaned.get('acceptance_threshold'),
                        cleaned.get('max_reviewers_per_proposal'),
                        cycle_id,
                    ],
                )
            messages.success(request, 'Grant Cycle updated successfully!')
            return redirect('cycle_list')
        messages.error(request, 'Please correct the grant cycle form and try again.')

    attach_cycle_display_fields(cycle)
    return render(request, 'chair/edit_cycle.html', {
        'cycle': cycle,
        'schools': schools
    })


def set_active_cycle(request, cycle_id):
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    if request.method != 'POST':
        return redirect('cycle_list')

    cycle = get_object_or_404(
        Grantcycles.objects.select_related('school'),
        cycle_id=cycle_id,
        school=chair.school,
    )

    set_active_cycle_for_school(chair.school_id, cycle.cycle_id)
    messages.success(request, f'"{cycle.cycle_name}" is now the active grant cycle for {chair.school.school_name}.')
    return redirect('cycle_list')


#p2

def proposal_list(request):
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    cycles = list(
        Grantcycles.objects
        .filter(school=chair.school)
        .order_by('-year', '-cycle_id')
    )

    selected_cycle_id = request.GET.get('cycle_id')

    proposals = (
        Proposals.objects
        .filter(department__school=chair.school)
        .select_related('department', 'department__school', 'pi_user', 'cycle')
        .order_by('-proposal_id')
    )

    if selected_cycle_id:
        proposals = proposals.filter(cycle_id=selected_cycle_id)

    return render(request, 'proposals/proposal_list.html', {
        'proposals': proposals,
        'cycles': cycles,
        'selected_cycle_id': str(selected_cycle_id) if selected_cycle_id else '',
    })


from .models import SrcChairs


def proposal_detail(request, proposal_id):
    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        return redirect('login')

    proposal = get_object_or_404(
        Proposals.objects.select_related('department', 'department__school'),
        proposal_id=proposal_id,
        department__school=chair.school,
    )
    docs = Proposaldocuments.objects.filter(proposal=proposal)

    # ✅ ADD THIS
    assignments = (
        Reviewassignments.objects
        .filter(proposal_id=proposal_id, is_active=True)
        .select_related("reviewer__user")
    )

    context = {
        'proposal': proposal,
        'docs': docs,
        'assignments': assignments,  # ✅ ADD THIS
        'is_chair': SrcChairs.objects.filter(user=request.user, is_active=True).exists()
    }
    return render(request, 'proposals/proposal_detail.html', context)


def assign_reviewer(request, proposal_id):
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        chair = SrcChairs.objects.select_related('school').get(user=request.user, is_active=True)
    except SrcChairs.DoesNotExist:
        messages.error(request, "Only SRC Chair can assign reviewers.")
        return redirect('login')

    try:
        proposal = Proposals.objects.select_related('department__school').get(
            proposal_id=proposal_id,
            department__school=chair.school,
        )
    except Proposals.DoesNotExist:
        messages.error(request, "Proposal not found.")
        return redirect('proposal_list')

    proposal_school = getattr(proposal.department, 'school', None)
    reviewers = (
        Reviewers.objects
        .filter(is_active=True, department__school=proposal_school)
        .exclude(user_id=proposal.pi_user_id)
        .select_related("user", "department", "department__school")
    )

    if request.method == "POST":
        reviewer_id = request.POST.get("reviewer")

        if not reviewer_id:
            messages.error(request, "Please select a reviewer.")
            return redirect('assign_reviewer', proposal_id=proposal_id)

        reviewer_obj = Reviewers.objects.filter(
            reviewer_id=reviewer_id,
            is_active=True,
            department__school=proposal_school
        ).first()
        if not reviewer_obj:
            messages.error(request, "Selected reviewer is not valid, not active, or not from the proposal's school.")
            return redirect('assign_reviewer', proposal_id=proposal_id)

        if reviewer_obj.user_id == proposal.pi_user_id:
            messages.error(request, "A reviewer cannot be assigned to review their own PI proposal.")
            return redirect('assign_reviewer', proposal_id=proposal_id)

        existing_assignment = Reviewassignments.objects.filter(
            proposal_id=proposal_id,
            reviewer_id=reviewer_id
        ).first()

        if existing_assignment and existing_assignment.is_active:
            email_sent, email_message = send_assignment_email(
                request=request,
                reviewer=reviewer_obj,
                proposal=proposal,
                assignment=existing_assignment,
            )
            if email_sent:
                messages.success(
                    request,
                    f"This reviewer is already assigned. The assignment email was resent to {reviewer_obj.user.email}.",
                )
            else:
                messages.warning(
                    request,
                    f"This reviewer is already assigned. {email_message}",
                )
            return redirect('assign_reviewer', proposal_id=proposal_id)

        current_count = Reviewassignments.objects.filter(proposal_id=proposal_id, is_active=True).count()
        cycle = proposal.cycle
        max_reviewers = cycle.max_reviewers_per_proposal if cycle and cycle.max_reviewers_per_proposal else 2
        if current_count >= max_reviewers:
            messages.error(request, f"Maximum of {max_reviewers} reviewers already assigned to this proposal.")
            return redirect('assign_reviewer', proposal_id=proposal_id)

        if existing_assignment:
            existing_assignment.assigned_at = timezone.now()
            existing_assignment.is_active = True
            existing_assignment.acceptance_status = 'PENDING'
            existing_assignment.save(update_fields=["assigned_at", "is_active", "acceptance_status"])
            assignment = existing_assignment
            messages.success(request, "Reviewer assignment reactivated successfully!")
        else:
            assignment = Reviewassignments.objects.create(
                proposal_id=proposal_id,
                reviewer_id=reviewer_id,
                assigned_at=timezone.now(),
                is_active=True,
                acceptance_status='PENDING'
            )
            messages.success(request, "Reviewer assigned successfully!")

        # Transition proposal to Under Stage 1 Review on first assignment
        if current_count == 0 and proposal.status == 'SUBMITTED':
            proposal.status = 'UNDER_STAGE_1_REVIEW'
            proposal.save(update_fields=['status'])

        email_sent, email_message = send_assignment_email(
            request=request,
            reviewer=reviewer_obj,
            proposal=proposal,
            assignment=assignment,
        )
        if email_sent:
            messages.success(request, email_message)
        else:
            messages.warning(request, email_message)

        return redirect('assign_reviewer', proposal_id=proposal_id)

    current_assignments = Reviewassignments.objects.filter(
        proposal_id=proposal_id,
        is_active=True
    ).select_related("reviewer__user")

    old_assignments = Reviewassignments.objects.filter(
        proposal_id=proposal_id,
        is_active=False
    ).select_related("reviewer__user")

    return render(request, 'proposals/assign_reviewer.html', {
        'proposal': proposal,
        'reviewers': reviewers,
        'current_assignments': current_assignments,
        'old_assignments': old_assignments
    })


def stage1_review(request, assignment_id):
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        assignment = Reviewassignments.objects.select_related("proposal").get(assignment_id=assignment_id)
    except Reviewassignments.DoesNotExist:
        messages.error(request, "Assignment not found.")
        return redirect('reviewer_dashboard')

    reviewer = Reviewers.objects.filter(user=request.user, is_active=True).first()
    if not reviewer:
        return redirect('login')

    if assignment.reviewer_id != reviewer.reviewer_id:
        messages.error(request, "You are not allowed to review this assignment.")
        return redirect('reviewer_dashboard')

    if Stage1Reviews.objects.filter(assignment=assignment).exists():
        messages.error(request, "Stage 1 review already submitted.")
        return redirect('reviewer_dashboard')

    # Enforce Stage 1 time window (compare naive local times)
    now = local_now_naive()
    def _naive(d):
        return d.replace(tzinfo=None) if d and getattr(d, 'tzinfo', None) else d
    cycle = assignment.proposal.cycle
    s1_start = _naive(cycle.stage1_start_date)
    s1_end   = _naive(cycle.stage1_end_date)
    if s1_start and now < s1_start:
        messages.error(request, f"Stage 1 review window has not opened yet. Opens: {s1_start.strftime('%b %d, %Y %H:%M:%S')}")
        return redirect('reviewer_dashboard')
    if s1_end and now > s1_end:
        messages.error(request, f"Stage 1 review window has closed. Closed: {s1_end.strftime('%b %d, %Y %H:%M:%S')}")
        return redirect('reviewer_dashboard')

    # Get proposal documents
    from django.db.models import Max
    proposal = assignment.proposal
    latest_ids = (
        Proposaldocuments.objects
        .filter(proposal=proposal)
        .values("document_type")
        .annotate(latest_id=Max("document_id"))
        .values_list("latest_id", flat=True)
    )
    docs = Proposaldocuments.objects.filter(document_id__in=latest_ids)

    if request.method == 'POST':
        try:
            score_originality = int(request.POST.get('score_originality'))
            score_clarity = int(request.POST.get('score_clarity'))
            score_lit_review = int(request.POST.get('score_lit_review'))
            score_methodology = int(request.POST.get('score_methodology'))
            score_impact = int(request.POST.get('score_impact'))
            score_publication = int(request.POST.get('score_publication'))
            score_budget = int(request.POST.get('score_budget'))
            score_timeframe = int(request.POST.get('score_timeframe'))
        except (TypeError, ValueError):
            messages.error(request, "All Stage 1 scores must be valid numbers.")
            return render(request, "reviews/stage1_review.html", {
                "assignment": assignment,
                "docs": docs,
            })

        narrative_comments = request.POST.get('narrative_comments')
        submitted_at = timezone.now()
        total_percentage = calculate_stage1_total(
            score_originality,
            score_clarity,
            score_lit_review,
            score_methodology,
            score_impact,
            score_publication,
            score_budget,
            score_timeframe,
        )

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO stage1reviews
                (assignment_id,
                 score_originality, score_clarity, score_lit_review, score_methodology,
                 score_impact, score_publication, score_budget, score_timeframe,
                 total_percentage, narrative_comments, is_submitted, submitted_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [
                assignment.assignment_id,
                score_originality, score_clarity, score_lit_review, score_methodology,
                score_impact, score_publication, score_budget, score_timeframe,
                total_percentage, narrative_comments, True, submitted_at
            ])

        messages.success(request, "Stage 1 review submitted!")
        return redirect('reviewer_dashboard')

    return render(request, "reviews/stage1_review.html", {
        "assignment": assignment,
        "docs": docs
    })

def stage2_review(request, assignment_id):
    assignment = get_object_or_404(Reviewassignments.objects.select_related("proposal"), pk=assignment_id)

    try:
        stage1 = Stage1Reviews.objects.get(assignment=assignment)
    except Stage1Reviews.DoesNotExist:
        messages.error(request, "Stage 1 must be submitted first.")
        return redirect("reviewer_dashboard")

    if Stage2Reviews.objects.filter(assignment=assignment).exists():
        messages.warning(request, "Stage 2 review already submitted.")
        return redirect("reviewer_dashboard")

    # Enforce Stage 2 time window (compare naive local times)
    now = local_now_naive()
    def _naive(d):
        return d.replace(tzinfo=None) if d and getattr(d, 'tzinfo', None) else d
    cycle = assignment.proposal.cycle
    s2_start = _naive(cycle.stage2_start_date)
    s2_end   = _naive(cycle.stage2_end_date)
    if s2_start and now < s2_start:
        messages.error(request, f"Stage 2 review window has not opened yet. Opens: {s2_start.strftime('%b %d, %Y %H:%M:%S')}")
        return redirect("reviewer_dashboard")
    if s2_end and now > s2_end:
        messages.error(request, f"Stage 2 review window has closed. Closed: {s2_end.strftime('%b %d, %Y %H:%M:%S')}")
        return redirect("reviewer_dashboard")

    # Move proposal into Under Stage 2 Review on first reviewer access
    proposal_obj = assignment.proposal
    if proposal_obj.status not in ('UNDER_STAGE_2_REVIEW', 'FINAL_ACCEPTED', 'FINAL_REJECTED'):
        proposal_obj.status = 'UNDER_STAGE_2_REVIEW'
        proposal_obj.save(update_fields=['status'])

    # Get latest docs
    from django.db.models import Max
    proposal = assignment.proposal
    latest_ids = (
        Proposaldocuments.objects
        .filter(proposal=proposal)
        .values("document_type")
        .annotate(latest_id=Max("document_id"))
        .values_list("latest_id", flat=True)
    )
    docs = Proposaldocuments.objects.filter(document_id__in=latest_ids)

    if request.method == "POST":
        concerns_addressed = request.POST.get("concerns_addressed")
        recommendation = request.POST.get("recommendation")
        revised_score = request.POST.get("revised_score")
        comments = request.POST.get("comments")

        if not all([concerns_addressed, recommendation, revised_score, comments]):
            messages.error(request, "All fields are required.")
            return render(request, "reviews/stage2_review.html", {
                "assignment": assignment
            })

        try:
            revised_score = int(revised_score)
        except ValueError:
            messages.error(request, "Revised score must be a number.")
            return render(request, "reviews/stage2_review.html", {
                "assignment": assignment
            })

        # ✅ Save exactly matching your DB structure
        Stage2Reviews.objects.create(
            assignment=assignment,
            concerns_addressed=concerns_addressed,
            recommendation=recommendation,
            revised_score=revised_score,
            comments=comments,
            submitted_at=timezone.now()
        )

        # 4️⃣ Update proposal status
        proposal = assignment.proposal

        if recommendation == "ACCEPT":
            proposal.status = "FINAL_ACCEPTED"
        elif recommendation == "REJECT":
            proposal.status = "FINAL_REJECTED"

        proposal.save()

        messages.success(request, "Stage 2 review submitted successfully.")
        return redirect("reviewer_dashboard")


    return render(request, "reviews/stage2_review.html", {
            "assignment": assignment,
            "docs": docs
        })


def stage1_review_result(request, assignment_id):
    if not request.user.is_authenticated:
        return redirect('login')

    assignment = get_object_or_404(
        Reviewassignments.objects.select_related("proposal", "reviewer__user"),
        assignment_id=assignment_id
    )
    is_chair = SrcChairs.objects.filter(user=request.user, is_active=True).exists()
    reviewer = Reviewers.objects.filter(user=request.user, is_active=True).first()

    if not is_chair and (not reviewer or assignment.reviewer_id != reviewer.reviewer_id):
        messages.error(request, "You are not allowed to view this Stage 1 result.")
        return redirect('dashboard')

    review = get_object_or_404(Stage1Reviews, assignment=assignment)
    ensure_stage1_total(review)
    back_url_name = 'reviewer_dashboard' if reviewer and assignment.reviewer_id == reviewer.reviewer_id else 'p3_chair_report_proposal'
    return render(request, "reviews/stage1_review_result.html", {
        "assignment": assignment,
        "review": review,
        "back_url_name": back_url_name,
    })


def stage2_review_result(request, assignment_id):
    if not request.user.is_authenticated:
        return redirect('login')

    assignment = get_object_or_404(
        Reviewassignments.objects.select_related("proposal", "reviewer__user"),
        assignment_id=assignment_id
    )
    is_chair = SrcChairs.objects.filter(user=request.user, is_active=True).exists()
    reviewer = Reviewers.objects.filter(user=request.user, is_active=True).first()

    if not is_chair and (not reviewer or assignment.reviewer_id != reviewer.reviewer_id):
        messages.error(request, "You are not allowed to view this Stage 2 result.")
        return redirect('dashboard')

    review = get_object_or_404(Stage2Reviews, assignment=assignment)
    back_url_name = 'reviewer_dashboard' if reviewer and assignment.reviewer_id == reviewer.reviewer_id else 'p3_chair_report_proposal'
    return render(request, "reviews/stage2_review_result.html", {
        "assignment": assignment,
        "review": review,
        "back_url_name": back_url_name,
    })


# 2. Add deactivate_assignment view
def deactivate_assignment(request, assignment_id):
    if not request.user.is_authenticated:
        return redirect('login')
    if not SrcChairs.objects.filter(user=request.user, is_active=True).exists():
        messages.error(request, "Only SRC Chair can deactivate reviewers.")
        return redirect('proposal_list')
    assignment = get_object_or_404(Reviewassignments, pk=assignment_id)
    assignment.is_active = False
    assignment.save(update_fields=["is_active"])
    messages.success(request, "Reviewer deactivated successfully.")
    return redirect('assign_reviewer', proposal_id=assignment.proposal_id)


def export_reviewers_excel_one_row(request):
    if not request.user.is_authenticated:
        return redirect("login")

    if not SrcChairs.objects.filter(user=request.user, is_active=True).exists():
        messages.error(request, "Only SRC Chair can export reviewer data.")
        return redirect("chair_reviewers")  # change if your view name differs

    now = timezone.now()

    # Current cycle = now between stage1_start_date and stage2_end_date
    cycle = (Grantcycles.objects
             .select_related("school")
             .filter(stage1_start_date__lte=now, stage2_end_date__gte=now)
             .order_by("-year", "-cycle_id")
             .first())

    # fallback: latest cycle
    if cycle is None:
        cycle = Grantcycles.objects.select_related("school").all().order_by("-year", "-cycle_id").first()

    if cycle is None:
        messages.error(request, "No grant cycle found.")
        return redirect("chair_reviewers")

    # All assignments for proposals in this cycle
    assignments = (
        Reviewassignments.objects
        .select_related(
            "reviewer__user",
            "reviewer__department",
            "proposal",
            "proposal__department",
        )
        .filter(proposal__cycle=cycle)
        .order_by("reviewer__user__username", "proposal__title")
    )

    # Group data: one row per reviewer
    grouped = {}
    for a in assignments:
        reviewer = a.reviewer
        if reviewer is None:
            continue

        rid = reviewer.reviewer_id
        user = reviewer.user

        if rid not in grouped:
            grouped[rid] = {
                "username": getattr(user, "username", ""),
                "email": getattr(user, "email", ""),
                "dept": getattr(getattr(reviewer, "department", None), "department_name", ""),
                "expertise": getattr(user, "area_of_expertise", ""),
                "max_load": getattr(reviewer, "max_review_load", ""),
                "reviewer_active": getattr(reviewer, "is_active", ""),
                "proposal_ids": [],
                "proposal_titles": [],
                "proposal_statuses": [],
                "assignment_ids": [],
                "assignment_statuses": [],
                "assignment_active_flags": [],
            }

        proposal = a.proposal
        grouped[rid]["proposal_ids"].append(str(getattr(proposal, "proposal_id", "")))
        grouped[rid]["proposal_titles"].append(str(getattr(proposal, "title", "")))
        grouped[rid]["proposal_statuses"].append(str(getattr(proposal, "status", "")))

        grouped[rid]["assignment_ids"].append(str(getattr(a, "assignment_id", "")))
        grouped[rid]["assignment_statuses"].append(str(getattr(a, "acceptance_status", "")))
        grouped[rid]["assignment_active_flags"].append(str(getattr(a, "is_active", "")))

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviewers (Current Cycle)"

    title = f"Reviewers & Assigned Proposals - {getattr(cycle,'cycle_name','')} ({getattr(cycle,'year','')})"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)

    headers = [
        "Cycle ID", "Cycle Name", "Cycle School", "Year",
        "Reviewer Username", "Reviewer Email", "Reviewer Department",
        "Expertise", "Max Review Load", "Reviewer Active",
        "Assigned Proposal IDs (comma)", "Assigned Proposal Titles (comma)", "Proposal Statuses (comma)",
        "Assignment IDs (comma)", "Assignment Statuses (comma)", "Assignment Active Flags (comma)",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = sorted(grouped.values(), key=lambda x: (x["username"] or "").lower())

    for r in rows:
        ws.append([
            getattr(cycle, "cycle_id", ""),
            getattr(cycle, "cycle_name", ""),
            getattr(getattr(cycle, "school", None), "school_name", ""),
            getattr(cycle, "year", ""),

            r["username"],
            r["email"],
            r["dept"],
            r["expertise"],
            r["max_load"],
            r["reviewer_active"],

            ", ".join([x for x in r["proposal_ids"] if x]),
            ", ".join([x for x in r["proposal_titles"] if x]),
            ", ".join([x for x in r["proposal_statuses"] if x]),

            ", ".join([x for x in r["assignment_ids"] if x]),
            ", ".join([x for x in r["assignment_statuses"] if x]),
            ", ".join([x for x in r["assignment_active_flags"] if x]),
        ])

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"

    # Auto column widths
    for col in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    filename = f"reviewers_current_cycle_{getattr(cycle,'cycle_id','cycle')}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
